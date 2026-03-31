from __future__ import annotations

from pathlib import Path

from .config import MAX_CONTENT_BYTES, SNIPPET_CHARS, SUPPORTED_TEXT_EXTENSIONS


def extract_text(file_path: Path) -> str:
    suffix = file_path.suffix.lower()
    try:
        if file_path.stat().st_size > MAX_CONTENT_BYTES:
            return ""
    except OSError:
        return ""

    try:
        if suffix in SUPPORTED_TEXT_EXTENSIONS:
            return file_path.read_text(encoding="utf-8", errors="ignore")[:SNIPPET_CHARS]
        if suffix == ".pdf":
            return _extract_pdf(file_path)
        if suffix == ".docx":
            return _extract_docx(file_path)
        if suffix == ".xlsx":
            return _extract_xlsx(file_path)
    except Exception:
        return ""
    return ""


def _extract_pdf(file_path: Path) -> str:
    from PyPDF2 import PdfReader

    reader = PdfReader(str(file_path))
    parts: list[str] = []
    for page in reader.pages[:5]:
        parts.append(page.extract_text() or "")
        if sum(len(part) for part in parts) >= SNIPPET_CHARS:
            break
    return "\n".join(parts)[:SNIPPET_CHARS]


def _extract_docx(file_path: Path) -> str:
    import docx

    document = docx.Document(str(file_path))
    text = "\n".join(paragraph.text for paragraph in document.paragraphs)
    return text[:SNIPPET_CHARS]


def _extract_xlsx(file_path: Path) -> str:
    from openpyxl import load_workbook

    workbook = load_workbook(filename=str(file_path), read_only=True, data_only=True)
    parts: list[str] = []
    for sheet in workbook.worksheets[:3]:
        for row in sheet.iter_rows(max_row=20, values_only=True):
            parts.append(" ".join(str(cell) for cell in row if cell is not None))
            if sum(len(part) for part in parts) >= SNIPPET_CHARS:
                return "\n".join(parts)[:SNIPPET_CHARS]
    return "\n".join(parts)[:SNIPPET_CHARS]
